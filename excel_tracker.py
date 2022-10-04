from openpyxl import Workbook,load_workbook

# Excel Structure
# Col A  - Name
# Col B - Date
# Col C - Month
# Col D - Year
# Col E - Email ID

def excel_extract(date, month):
    wb = load_workbook('birthday.xlsx')
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    final_list = []

    for row in range(1,max_row+1):
        if ws['B'+str(row)].value == date and ws['C'+str(row)].value == month:
            curr_dict = {"name": ws["A" + str(row)].value, "email_id": ws["E" + str(row)].value}
            final_list.append(curr_dict)

    return final_list


